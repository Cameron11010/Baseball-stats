import seaborn as sns
from KNearestNeighbour import *
from openpyxl import load_workbook, Workbook

dataset = load_workbook(filename="stats.xlsx")

file_path = "stats_norm.xlsx"

train = dataset['Train']

test = dataset['Test']

validation = dataset['Validation']

headings = [train.cell(row=1, column=i).value for i in range(1, train.max_column + 1)]

test_headings = [test.cell(row=1, column=i).value for i in range(1, test.max_column + 1)]

validation_headings = [validation.cell(row=1, column=i).value for i in range(1, validation.max_column + 1)]


def sheet_exists(file_path, sheet_name):
    try:
        workbook = load_workbook(filename=file_path)
        return sheet_name in workbook.sheetnames
    except FileNotFoundError:
        return False


def copy_sheet(source_sheet_name, target_sheet_name, target_file_path):
    try:
        # Load the existing workbook
        target_workbook = load_workbook(filename=target_file_path)
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        target_workbook = Workbook()

    # Create the target sheet
    if target_sheet_name in target_workbook.sheetnames:
        # If the sheet already exists, remove it
        target_workbook.remove(target_workbook[target_sheet_name])
    target_sheet = target_workbook.create_sheet(title=target_sheet_name)

    # Load the source sheet
    source_sheet = dataset[source_sheet_name]

    # Copy contents from source to target sheet
    for row in source_sheet.iter_rows(values_only=True):
        target_sheet.append(row)

    # Save the workbook
    target_workbook.save(filename=target_file_path)


def normalise(sheet):
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row contains headers
        if not all(val is None for val in row):  # Skip over rows with all null values
            data.append(list(row))

    # Specify columns for normalization
    columns_to_normalize = [(7, 12), (14, 19), (21, 26), (28, 33), (35, 40), (42, 47), (49, 54), (56, 61), (63, 68)]

    # Apply min-max normalization to specified columns
    for start, end in columns_to_normalize:
        for row_data in data:
            col_values = [val for val in row_data[start - 1:end] if val is not None]  # Exclude null values
            if col_values:  # Skip over if all values are null
                min_val = min(col_values)
                max_val = max(col_values)
                for i in range(start - 1, end):
                    if row_data[i] is not None:  # Skip over null values
                        row_data[i] = (row_data[i] - min_val) / (max_val - min_val)

    return data


sheet_name = "Norm_Train"
if not sheet_exists(file_path, sheet_name):
    source_sheet_name = "Train"
    target_sheet_name = "Norm_Train"
    copy_sheet(source_sheet_name, target_sheet_name, file_path)
norm_workbook = load_workbook(filename=file_path)
norm_sheet = norm_workbook[sheet_name]
norm_data = normalise(norm_sheet)

for row_idx, row_data in enumerate(norm_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        norm_sheet.cell(row=row_idx, column=col_idx, value=value)

norm_workbook.save(filename=file_path)

sheet_name = "Norm_Test"
if not sheet_exists(file_path, sheet_name):
    source_sheet_name = "Test"
    target_sheet_name = "Norm_Test"
    copy_sheet(source_sheet_name, target_sheet_name, file_path)
norm_workbook = load_workbook(filename=file_path)
norm_sheet = norm_workbook[sheet_name]
norm_data = normalise(norm_sheet)

for row_idx, row_data in enumerate(norm_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        norm_sheet.cell(row=row_idx, column=col_idx, value=value)

norm_workbook.save(filename=file_path)

sheet_name = "Norm_Validation"
if not sheet_exists(file_path, sheet_name):
    source_sheet_name = "Validation"
    target_sheet_name = "Norm_Validation"
    copy_sheet(source_sheet_name, target_sheet_name, file_path)
norm_workbook = load_workbook(filename=file_path)
norm_sheet = norm_workbook[sheet_name]
norm_data = normalise(norm_sheet)

for row_idx, row_data in enumerate(norm_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        norm_sheet.cell(row=row_idx, column=col_idx, value=value)

norm_workbook.save(filename=file_path)

norm_dataset = load_workbook(filename="stats_norm.xlsx")

train = norm_dataset['Norm_Train']
test = norm_dataset['Norm_Test']
validation = norm_dataset['Norm_Validation']

pitch_ranges = {
    "fseam": ([7, 12]),
    "slider": (14, 19),
    "changeup": (21, 26),
    "curve": (28, 33),
    "sinker": (35, 40),
    "cutter": (42, 47),
    "splitter": (49, 54),
    "knuckle": (56, 61),
    "tseam": (63, 68),
}

results = {}

for pitch, pitch_range in pitch_ranges.items():
    results[pitch] = []
    for i in range(2, train.max_row + 1):
        row_data = []
        for j in range(pitch_range[0], pitch_range[1] + 1):
            cell_value = train.cell(row=i, column=j).value
            if cell_value is not None or cell_value == 0:
                row_data.append(cell_value)
        if row_data:  # Only append non-empty rows
            results[pitch].append(row_data)

fseam = results["fseam"]
slider = results["slider"]
changeup = results["changeup"]
curve = results["curve"]
sinker = results["sinker"]
cutter = results["cutter"]
splitter = results["splitter"]
knuckle = results["knuckle"]
tseam = results["tseam"]

points = {'4-seam fastball': fseam,
          'sinker': sinker,
          'cutter': cutter,
          '2-seam fastball': tseam,
          'splitter': splitter,
          'changeup': changeup,
          'slider': slider,
          'curve': curve,
          'knuckle': knuckle,
          }


def pitch_classification(column_numbers):
    if column_numbers == [7, 8, 9, 10, 11, 12]:
        return "4-seam fastball"
    elif column_numbers == [14, 15, 16, 17, 18, 19]:
        return "slider"
    if column_numbers == [21, 22, 23, 24, 25, 26]:
        return "changeup"
    if column_numbers == [28, 29, 30, 31, 32, 33]:
        return "curve"
    if column_numbers == [35, 36, 37, 38, 39, 40]:
        return "sinker"
    if column_numbers == [42, 43, 44, 45, 46, 47]:
        return "cutter"
    if column_numbers == [49, 50, 51, 52, 53, 54]:
        return "splitter"
    if column_numbers == [56, 57, 58, 59, 60, 61]:
        return "knuckle"
    if column_numbers == [63, 64, 65, 66, 67, 68]:
        return "2-seam fastball"
    else:
        return "Unknown pitch type"


def plot_confusion_matrix(tp, tn, fp, fn, title, labels=None):
    confusion_matrix = np.array([[tp, fp], [fn, tn]])

    if labels is None:
        labels = ['Positive', 'Negative']

    sns.heatmap(confusion_matrix, annot=True, cmap="Blues", fmt="d", xticklabels=labels, yticklabels=labels)
    plt.xlabel('Predicted Labels')
    plt.ylabel('True Labels')
    plt.title(title)
    plt.show()


ranges = [(7, 12), (14, 19), (21, 26), (28, 33), (35, 40), (42, 47), (49, 54), (56, 61), (63, 68)]


unknown_pitch = []

for start, end in ranges:
    for j in range(2, test.max_row + 1):
        column_values = []  # Reset column_values for each row
        for start, end in ranges:
            for i in range(start, end + 1):
                cell_value = test.cell(row=j, column=i).value
                if cell_value is not None or cell_value == 0:
                    column_values.append((cell_value, i))
        if column_values:
            unknown_pitch.extend(column_values)

unknown_pitch_val = []
for start, end in ranges:
    for j in range(2, validation.max_row + 1):
        column_values = []  # Reset column_values for each row
        for start, end in ranges:
            for i in range(start, end + 1):
                cell_value = validation.cell(row=j, column=i).value
                if cell_value is not None or cell_value == 0:
                    column_values.append((cell_value, i))
        if column_values:
            unknown_pitch_val.extend(column_values)

unknown_pitch = unknown_pitch_val
clf = KNearestNeighbours()
clf.fit(points)

val_test = True
while val_test:
    display = False
    ans = input("Would you like to see the 3D graphs?(y/n)")
    if ans == "y":
        display = True
    clf = KNearestNeighbours()
    clf.fit(points)

    total_predictions = 0
    correct_predictions = 0

    # Separate each new unknown pitch into variables for each set of 6 features
    num_variables = len(unknown_pitch) // (6 * len(ranges))
    print("unknown P len ", (len(unknown_pitch) / 6), "\nranges len:", len(ranges), "\nnum variable", num_variables)

    for i in range(num_variables):
        globals()[f'unknown_pitch {i + 1}'] = unknown_pitch[i * 6: (i + 1) * 6]

    true_pos_fourseam = 0
    true_neg_fourseam = 0
    false_pos_fourseam = 0
    false_neg_fourseam = 0

    true_pos_slider = 0
    true_neg_slider = 0
    false_pos_slider = 0
    false_neg_slider = 0

    true_pos_changeup = 0
    true_neg_changeup = 0
    false_pos_changeup = 0
    false_neg_changeup = 0

    true_pos_curve = 0
    true_neg_curve = 0
    false_pos_curve = 0
    false_neg_curve = 0

    true_pos_sinker = 0
    true_neg_sinker = 0
    false_pos_sinker = 0
    false_neg_sinker = 0

    true_pos_cutter = 0
    true_neg_cutter = 0
    false_pos_cutter = 0
    false_neg_cutter = 0

    true_pos_splitter = 0
    true_neg_splitter = 0
    false_pos_splitter = 0
    false_neg_splitter = 0

    true_pos_knuckle = 0
    true_neg_knuckle = 0
    false_pos_knuckle = 0
    false_neg_knuckle = 0

    true_pos_twoseam = 0
    true_neg_twoseam = 0
    false_pos_twoseam = 0
    false_neg_twoseam = 0

    # Print the variables
    for i in range(num_variables):
        print(f'unknown_pitch {i + 1}: {globals()[f"unknown_pitch {i + 1}"]}')
        guess_pitch = [tup[0] for tup in globals()[f"unknown_pitch {i + 1}"]]
        guess_pitch_types = [test_headings[tup[1] - 7] for tup in globals()[f"unknown_pitch {i + 1}"]]
        guess_pitch_column_numbers = [tup[1] for tup in globals()[f"unknown_pitch {i + 1}"]]

        correct_pitch = pitch_classification(guess_pitch_column_numbers)
        prediction = clf.predict(guess_pitch, i + 1, display)

        print("Category prediction:", prediction)
        print("Correct pitch:", correct_pitch)

        if prediction == "4-seam fastball" and correct_pitch == "4-seam fastball":
            true_pos_fourseam += 1
        if prediction == "4-seam fastball" and correct_pitch != "4-seam fastball":
            false_pos_fourseam += 1
        if prediction != "4-seam fastball" and correct_pitch == "4-seam fastball":
            false_neg_fourseam += 1
        if prediction != "4-seam fastball" and correct_pitch != "4-seam fastball":
            true_neg_fourseam += 1

        if prediction == "slider" and correct_pitch == "slider":
            true_pos_slider += 1
        if prediction == "slider" and correct_pitch != "slider":
            false_pos_slider += 1
        if prediction != "slider" and correct_pitch == "slider":
            false_neg_slider += 1
        if prediction != "slider" and correct_pitch != "slider":
            true_neg_slider += 1

        if prediction == "changeup" and correct_pitch == "changeup":
            true_pos_changeup += 1
        if prediction == "changeup" and correct_pitch != "changeup":
            false_pos_changeup += 1
        if prediction != "changeup" and correct_pitch == "changeup":
            false_neg_changeup += 1
        if prediction != "changeup" and correct_pitch != "changeup":
            true_neg_changeup += 1

        if prediction == "curve" and correct_pitch == "curve":
            true_pos_curve += 1
        if prediction == "curve" and correct_pitch != "curve":
            false_pos_curve += 1
        if prediction != "curve" and correct_pitch == "curve":
            false_neg_curve += 1
        if prediction != "curve" and correct_pitch != "curve":
            true_neg_curve += 1

        if prediction == "sinker" and correct_pitch == "sinker":
            true_pos_sinker += 1
        if prediction == "sinker" and correct_pitch != "sinker":
            false_pos_sinker += 1
        if prediction != "sinker" and correct_pitch == "sinker":
            false_neg_sinker += 1
        if prediction != "sinker" and correct_pitch != "sinker":
            true_neg_sinker += 1

        if prediction == "cutter" and correct_pitch == "cutter":
            true_pos_cutter += 1
        if prediction == "cutter" and correct_pitch != "cutter":
            false_pos_cutter += 1
        if prediction != "cutter" and correct_pitch == "cutter":
            false_neg_cutter += 1
        if prediction != "cutter" and correct_pitch != "cutter":
            true_neg_cutter += 1

        if prediction == "splitter" and correct_pitch == "splitter":
            true_pos_splitter += 1
        if prediction == "splitter" and correct_pitch != "splitter":
            false_pos_splitter += 1
        if prediction != "splitter" and correct_pitch == "splitter":
            false_neg_splitter += 1
        if prediction != "splitter" and correct_pitch != "splitter":
            true_neg_splitter += 1

        if prediction == "knuckle" and correct_pitch == "knuckle":
            true_pos_knuckle += 1
        if prediction == "knuckle" and correct_pitch != "knuckle":
            false_pos_knuckle += 1
        if prediction != "knuckle" and correct_pitch == "knuckle":
            false_neg_knuckle += 1
        if prediction != "knuckle" and correct_pitch != "knuckle":
            true_neg_knuckle += 1

        if prediction == "2-seam fastball" and correct_pitch == "2-seam fastball":
            true_pos_twoseam += 1
        if prediction == "2-seam fastball" and correct_pitch != "2-seam fastball":
            false_pos_twoseam += 1
        if prediction != "2-seam fastball" and correct_pitch == "2-seam fastball":
            false_neg_twoseam += 1
        if prediction != "2-seam fastball" and correct_pitch != "2-seam fastball":
            true_neg_twoseam += 1

        if prediction == correct_pitch:
            correct_predictions += 1
        else:
            correct_predictions = correct_predictions
            plt.show()

        total_predictions += 1

    print("\n4-seam fastballs: \ntrue pos: ", true_pos_fourseam, " true neg: ", true_neg_fourseam)
    print("false pos: ", false_pos_fourseam, " false neg: ", false_neg_fourseam)
    plot_confusion_matrix(true_pos_fourseam, true_neg_fourseam, false_pos_fourseam, false_neg_fourseam,
                          "4-seam confusion matrix")

    print("\nSliders: \ntrue pos: ", true_pos_slider, " true neg: ", true_neg_slider)
    print("false pos: ", false_pos_slider, " false neg: ", false_neg_slider)
    plot_confusion_matrix(true_pos_slider, true_neg_slider, false_pos_slider, false_neg_slider, "slider confusion matrix")

    print("\nChangeups: \ntrue pos: ", true_pos_changeup, " true neg: ", true_neg_changeup)
    print("false pos: ", false_pos_changeup, " false neg: ", false_neg_changeup)
    plot_confusion_matrix(true_pos_changeup, true_neg_changeup, false_pos_changeup, false_neg_changeup,
                          "changeup confusion matrix")

    print("\nCurves: \ntrue pos: ", true_pos_curve, " true neg: ", true_neg_curve)
    print("false pos: ", false_pos_curve, " false neg: ", false_neg_curve)
    plot_confusion_matrix(true_pos_curve, true_neg_curve, false_pos_curve, false_neg_curve, "curve confusion matrix")

    print("\nSinkers: \ntrue pos: ", true_pos_sinker, " true neg: ", true_neg_sinker)
    print("false pos: ", false_pos_sinker, " false neg: ", false_neg_sinker)
    plot_confusion_matrix(true_pos_sinker, true_neg_sinker, false_pos_sinker, false_neg_sinker, "sinker confusion matrix")

    print("\nCutters: \ntrue pos: ", true_pos_cutter, " true neg: ", true_neg_cutter)
    print("false pos: ", false_pos_cutter, " false neg: ", false_neg_cutter)
    plot_confusion_matrix(true_pos_cutter, true_neg_cutter, false_pos_cutter, false_neg_cutter, "cutter confusion matrix")

    print("\nSplitters: \ntrue pos: ", true_pos_splitter, " true neg: ", true_neg_splitter)
    print("false pos: ", false_pos_splitter, " false neg: ", false_neg_splitter)
    plot_confusion_matrix(true_pos_splitter, true_neg_splitter, false_pos_splitter, false_neg_splitter,
                          "splitter confusion matrix")

    print("\nKnuckles: \ntrue pos: ", true_pos_knuckle, " true neg: ", true_neg_knuckle)
    print("false pos: ", false_pos_knuckle, " false neg: ", false_neg_knuckle)
    plot_confusion_matrix(true_pos_knuckle, true_neg_knuckle, false_pos_knuckle, false_neg_knuckle,
                          "knuckle confusion matrix")

    print("\n2-seam fastballs: \ntrue pos: ", true_pos_twoseam, " true neg: ", true_neg_twoseam)
    print("false pos: ", false_pos_twoseam, " false neg", false_neg_twoseam)
    plot_confusion_matrix(true_pos_twoseam, true_neg_twoseam, false_pos_twoseam, false_neg_twoseam,
                          "2-seam confusion matrix")

    accuracy = (correct_predictions / total_predictions) * 100

    print("\nCorrect count: ", correct_predictions)
    print("Total count: ", total_predictions)

    print("Accuracy = ", accuracy)

    validation_test_run = False
    val_test = input("Run Validation test?(y/n)")
    if val_test == 'y' and not validation_test_run:
        unknown_pitch = unknown_pitch_val
        clf = KNearestNeighbours()
        clf.fit(points)
        validation_test_run = True
    else:
        break
